import io
import streamlit as st
import pdfplumber
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side

# ============================================================
# ★ 調整パラメータ（フォーマットが変わった場合はここを変更）
# ============================================================
QTY_ROWS_PER_PAGE = 15   # 数量書：1ページあたりのデータ行数
CAT_ROWS_PER_PAGE = 14   # 工事別数量書：1ページあたりのデータ行数
SUBTOTAL_NAMES = {'小 計', '小　計', '計', '合 計', '合　計', '合　　計', '総 合 計'}
SKIP_VALS      = {'数 量 書', '工 種 別 数 量 書', '総 括 数 量 書', '種 別', '内 訳', '工 種',
                  '数    量    書', '工 種 別 内 訳 書', '総　括　書', '数量書', '工種別数量書'}
SUMMARY_KEEP   = {'工事価格 合計', '工事価格　合計', '消費税額', '総 合 計', '総　合　計'}
# ============================================================

# ---------- ユーティリティ ----------

def thin_border():
    s = Side(style='thin')
    return Border(left=s, right=s, top=s, bottom=s)

def border_merged(ws, cell_range):
    """結合範囲の全構成セルに四辺罫線を付与（環境差で枠線が消えるのを防ぐ）。
    cell_range 例: 'A1:E1'"""
    from openpyxl.utils import range_boundaries
    min_c, min_r, max_c, max_r = range_boundaries(cell_range)
    for rr in range(min_r, max_r + 1):
        for cc in range(min_c, max_c + 1):
            ws.cell(rr, cc).border = thin_border()

def finalize_borders(ws, ncols, no_border_titles=None):
    """データ範囲の全セル（結合セル内側含む）に四辺罫線を一括付与。
    結合セルは外周セルにも外向き罫線を明示設定し、環境差で枠線が
    消えるのを防ぐ。ページ番号行・枠外タイトル行は除外。"""
    no_border_titles = no_border_titles or set()

    def skip_row(r):
        av = ws.cell(r, 1).value
        sval = av.strip() if isinstance(av, str) else ''
        if sval and is_page_num(sval):
            return True
        if sval and sval in no_border_titles:
            return True
        return False

    # 1) まず全セルに四辺罫線
    for r in range(1, ws.max_row + 1):
        if skip_row(r):
            continue
        for c in range(1, ncols + 1):
            ws.cell(r, c).border = thin_border()

    # 2) 結合セルの外周セルに、外向きの罫線を明示設定（内側の縦/横線が
    #    保存時に消えても外枠は残るようにする）
    s = Side(style='thin')
    for m in list(ws.merged_cells.ranges):
        r1, r2, c1, c2 = m.min_row, m.max_row, m.min_col, m.max_col
        if skip_row(r1):
            continue
        for r in range(r1, r2 + 1):
            for c in range(c1, c2 + 1):
                cell = ws.cell(r, c)
                b = cell.border
                top    = s if r == r1 else b.top
                bottom = s if r == r2 else b.bottom
                left   = s if c == c1 else b.left
                right  = s if c == c2 else b.right
                cell.border = Border(top=top, bottom=bottom, left=left, right=right)

def apply_cell(cell, value, bold=False, center=False, right=False):
    cell.value = value
    cell.font = Font(name='MS Gothic', size=9, bold=bold)
    halign = 'right' if right else ('center' if center else 'left')
    cell.alignment = Alignment(horizontal=halign, vertical='center', wrap_text=True)
    cell.border = thin_border()

def set_formula(cell, formula, right=False):
    cell.value = formula
    cell.font = Font(name='MS Gothic', size=9)
    cell.alignment = Alignment(horizontal='right' if right else 'center', vertical='center')
    cell.border = thin_border()

def clean_qty(val):
    try:
        f = float(val)
        return int(f) if f == int(f) else f
    except:
        return val

def is_page_num(val):
    if not val: return False
    parts = val.replace(' ', '').split('/')
    return len(parts) == 2 and all(p.isdigit() for p in parts)

# ---------- PDF解析 ----------

def extract_all_pages(pdf_file, progress_cb=None):
    summary_rows, category_rows, quantity_rows = [], [], []
    with pdfplumber.open(pdf_file) as pdf:
        total = len(pdf.pages)
        for i, page in enumerate(pdf.pages):
            if progress_cb:
                progress_cb(i / total, f'{i+1} / {total} ページ解析中…')
            text  = page.extract_text() or ''
            first = text.split('\n')[0].strip()
            if   '総' in first and '括' in first: ptype = 'summary'
            elif '工' in first and '種' in first: ptype = 'category'
            elif '数' in first and '量' in first: ptype = 'quantity'
            else: continue
            tables = page.extract_tables()
            if not tables: continue
            after_header = False  # 「内訳/種別」ヘッダー直後フラグ（工事全体名行の検出用）
            first_in_page = True  # このPDFページで最初に追加するデータ行か
            def tag_first(row_list):
                # PDFページ境界マーカー：このページの最初のデータ行に True を付ける
                nonlocal first_in_page
                row_list.append('__PAGESTART__' if first_in_page else '')
                first_in_page = False
            for row in tables[0]:
                if not row: continue
                cleaned = [(str(c).replace('\n', ' ').strip() if c else '') for c in row]
                # ページ番号行スキップ
                if is_page_num(cleaned[0]): continue
                if is_page_num(cleaned[1] if len(cleaned) > 1 else ''): continue
                # 内訳/種別ヘッダー検出（SKIP_VALS判定より前にフラグを立てる）
                hdr_set = {'内 訳', '内　訳'} if ptype in ('summary', 'quantity') else {'種 別', '種　別'}
                if cleaned[0] in hdr_set or (len(cleaned) > 1 and cleaned[1] in hdr_set):
                    after_header = True
                    continue
                # ヘッダー直後の「番号なし・名称あり・内容なし」＝工事全体名行 → 見出しとして残す
                if (after_header and cleaned[1] and not cleaned[0]
                        and (len(cleaned) < 3 or not cleaned[2])):
                    after_header = False
                    if   ptype == 'summary':  summary_rows.append(cleaned)
                    elif ptype == 'category': category_rows.append(cleaned)
                    elif ptype == 'quantity':
                        tag_first(cleaned); quantity_rows.append(cleaned)
                    continue
                after_header = False
                # スキップ対象文字列
                if cleaned[0] in SKIP_VALS: continue
                if len(cleaned) > 1 and cleaned[1] in SKIP_VALS: continue
                if ptype == 'summary':
                    if (cleaned[1] and not cleaned[0] and not cleaned[2]
                            and cleaned[1] not in SUMMARY_KEEP): continue
                if ptype == 'category':
                    if cleaned[1] in {'種 別', '種　別'}: continue
                    if (cleaned[1] and not cleaned[0] and not cleaned[2]
                            and cleaned[1] not in SUBTOTAL_NAMES): continue
                if   ptype == 'summary':  summary_rows.append(cleaned)
                elif ptype == 'category': category_rows.append(cleaned)
                elif ptype == 'quantity':
                    tag_first(cleaned); quantity_rows.append(cleaned)
    return summary_rows, category_rows, quantity_rows

# ---------- 数量書 ----------

def write_qty_page_break(ws, r, page_num, total_pages):
    """数量書のページ区切り（空行→ページ番号→タイトル→ヘッダー2行→空行）"""
    # 空行（前ページ末尾の空行：データ行と同じ高さ25、全列に罫線）
    ws.merge_cells(f'B{r}:H{r}')
    for c in range(1, 9):
        ws.cell(r,c).border = thin_border()
    ws.row_dimensions[r].height = 25
    r += 1
    # ページ番号
    ws.merge_cells(f'A{r}:H{r}')
    ws.cell(r,1).value = f' {page_num} / {total_pages}'
    ws.cell(r,1).font  = Font(name='MS Gothic', size=8)
    ws.row_dimensions[r].height = 15
    r += 1
    # タイトル
    ws.merge_cells(f'A{r}:H{r}')
    ws.cell(r,1).value = '数    量    書'
    ws.cell(r,1).font  = Font(name='MS Gothic', size=11, bold=True)
    ws.cell(r,1).alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[r].height = 20
    r += 1
    # ヘッダー1行目（A-B結合でAB間の縦線を消す）
    h1 = r
    ws.merge_cells(f'A{h1}:B{h1}')
    ws.cell(h1,1).value = '種　　別'
    ws.cell(h1,1).font  = Font(name='MS Gothic', size=9)
    ws.cell(h1,1).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(h1,1).border = thin_border()
    for col, h in zip([3,4,5,6,7,8], ['形状・寸法','数  量','単位','単価(円)','金　額(円)','摘      要']):
        ws.cell(h1,col).value = h
        ws.cell(h1,col).font  = Font(name='MS Gothic', size=9)
        ws.cell(h1,col).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.cell(h1,col).border = thin_border()
    ws.row_dimensions[h1].height = 19.5
    r += 1
    # ヘッダー2行目（A-B結合）
    h2 = r
    ws.merge_cells(f'A{h2}:B{h2}')
    ws.cell(h2,1).value = '内　　訳'
    ws.cell(h2,1).font  = Font(name='MS Gothic', size=9)
    ws.cell(h2,1).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(h2,1).border = thin_border()
    for col in range(3, 9):
        ws.merge_cells(f'{chr(64+col)}{h1}:{chr(64+col)}{h2}')
    ws.row_dimensions[h2].height = 19.5
    r += 1
    # 空行
    ws.merge_cells(f'B{r}:H{r}')
    for c in range(1, 9):
        ws.cell(r,c).border = thin_border()
    ws.row_dimensions[r].height = 25
    r += 1
    return r

def build_quantity_sheet(ws, rows):
    ws.title = '数量書'
    for col, w in zip('ABCDEFGH', [8, 32, 38, 10, 8, 14, 14, 28]):
        ws.column_dimensions[col].width = w

    # 空データ行は除外するが、__PAGESTART__マーカーは次の有効行へ引き継ぐ
    data = []
    pending_marker = False
    for row in rows:
        if len(row) >= 9 and row[8] == '__PAGESTART__':
            pending_marker = True
        if len(row) >= 2 and not all(v == '' for v in row[:8]):
            new_row = list(row[:8])
            new_row.append('__PAGESTART__' if pending_marker else '')
            data.append(new_row)
            pending_marker = False
    # PDFの実ページ数を数える（元rowsの__PAGESTART__マーカーの数）
    total_pages = sum(1 for row in rows if len(row) >= 9 and row[8] == '__PAGESTART__')
    if total_pages < 1:
        total_pages = 1

    # ── ヘッダー（1ページ目） ──
    r = 1
    ws.merge_cells(f'A{r}:H{r}')
    ws.cell(r,1).value = '数    量    書'
    ws.cell(r,1).font  = Font(name='MS Gothic', size=11, bold=True)
    ws.cell(r,1).alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[r].height = 20
    r += 1

    h1 = r
    ws.merge_cells(f'A{h1}:B{h1}')
    ws.cell(h1,1).value = '種　　別'
    ws.cell(h1,1).font  = Font(name='MS Gothic', size=9)
    ws.cell(h1,1).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(h1,1).border = thin_border()
    for col, h in zip([3,4,5,6,7,8], ['形状・寸法','数  量','単位','単価(円)','金　額(円)','摘      要']):
        ws.cell(h1,col).value = h
        ws.cell(h1,col).font  = Font(name='MS Gothic', size=9)
        ws.cell(h1,col).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.cell(h1,col).border = thin_border()
    ws.row_dimensions[h1].height = 19.5
    r += 1

    h2 = r
    ws.merge_cells(f'A{h2}:B{h2}')
    ws.cell(h2,1).value = '内　　訳'
    ws.cell(h2,1).font  = Font(name='MS Gothic', size=9)
    ws.cell(h2,1).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(h2,1).border = thin_border()
    for col in range(3, 9):
        ws.merge_cells(f'{chr(64+col)}{h1}:{chr(64+col)}{h2}')
    ws.row_dimensions[h2].height = 19.5
    r += 1

    # ── データ行（PDFの実ページ境界で改ページ：忠実再現） ──
    subtotal_map      = {}
    current_big       = None
    current_sub       = None
    section_data_rows = []
    page_num          = 1
    first_data_done   = False  # 最初のデータ行を出力済みか

    for raw in data:
        page_start = len(raw) >= 9 and raw[8] == '__PAGESTART__'
        vals = (list(raw[:8]) + [''] * 8)[:8]
        num, name, shape, qty, unit, price, amt, note = vals
        bold        = bool(num)
        is_subtotal = name in SUBTOTAL_NAMES

        # PDFページが変わったら改ページ（ただし最初のページ先頭では入れない）
        if page_start and first_data_done:
            r = write_qty_page_break(ws, r, page_num, total_pages)
            page_num += 1
        first_data_done = True

        # 工事全体名行（番号なし・名称あり・数量なし、小計系でない）＝見出し行：B~Hを結合
        if name and not num and not qty and not is_subtotal:
            apply_cell(ws.cell(r,1), '', center=True)
            apply_cell(ws.cell(r,2), name, bold=True)
            for c in range(3, 9):
                ws.cell(r,c).border = thin_border()
            ws.merge_cells(f'B{r}:H{r}')
            ws.row_dimensions[r].height = 25
            r += 1
            continue

        if num.startswith('【'):
            current_big = num; current_sub = None; section_data_rows = []
        elif num:
            current_sub = num; section_data_rows = []

        apply_cell(ws.cell(r,1), num,   bold=bold, center=True)
        apply_cell(ws.cell(r,2), name,  bold=bold or is_subtotal)
        apply_cell(ws.cell(r,3), shape)
        apply_cell(ws.cell(r,4), clean_qty(qty) if qty else '', center=True)
        apply_cell(ws.cell(r,5), unit,  center=True)
        apply_cell(ws.cell(r,6), '',    center=True)

        if qty and unit and not is_subtotal and not bold:
            set_formula(ws.cell(r,7), f'=INT(F{r}*D{r})')
            section_data_rows.append(r)
        elif is_subtotal:
            # 改ページ行を挟むため加算式で連結（連続範囲SUMだとゴミ行を巻き込む）
            if section_data_rows:
                set_formula(ws.cell(r,7),
                            '=' + '+'.join(f'G{x}' for x in section_data_rows))
            else:
                apply_cell(ws.cell(r,7), '', center=True)
            subtotal_map[(current_big, current_sub)] = r
            section_data_rows = []
        else:
            apply_cell(ws.cell(r,7), '', center=True)

        apply_cell(ws.cell(r,8), note)
        ws.row_dimensions[r].height = 25  # ★ 全データ行25pt均一
        r += 1

    # 罫線を一括補完（結合セル内側含む）。数量書のタイトルは枠外なので除外
    finalize_borders(ws, 8, no_border_titles={'数    量    書'})
    return subtotal_map

# ---------- 工事別数量書 ----------

def write_cat_header_block(ws, r, page_num, total_pages):
    """工事別数量書のページ区切り（空行→ページ番号→タイトル→ヘッダー2行→空行）"""
    # 空行（前ページ末尾の空行：データ行と同じ高さ25、全列に罫線）
    ws.merge_cells(f'B{r}:E{r}')
    for c in range(1, 6):
        ws.cell(r,c).border = thin_border()
    ws.row_dimensions[r].height = 25
    r += 1
    # ページ番号
    ws.merge_cells(f'A{r}:E{r}')
    ws.cell(r,1).value = f' {page_num} / {total_pages}'
    ws.cell(r,1).font  = Font(name='MS Gothic', size=8)
    ws.row_dimensions[r].height = 15
    r += 1
    # タイトル（中央寄せ＋外枠罫線）
    ws.merge_cells(f'A{r}:E{r}')
    ws.cell(r,1).value = '工 種 別 数 量 書'
    ws.cell(r,1).font  = Font(name='MS Gothic', size=10, bold=True)
    ws.cell(r,1).alignment = Alignment(horizontal='center', vertical='center')
    for c in range(1, 6):
        ws.cell(r,c).border = thin_border()
    ws.row_dimensions[r].height = 20
    r += 1
    # ヘッダー1行目（A-B結合で「工種」とし、A-B間の縦線を消す）
    h1 = r
    ws.merge_cells(f'A{h1}:B{h1}')
    ws.cell(h1,1).value = '工    種'
    ws.cell(h1,1).font  = Font(name='MS Gothic', size=9)
    ws.cell(h1,1).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(h1,1).border = thin_border()
    for col, h in [(3,'内  容 (数量)'),(4,'金　額 （円）'),(5,' 摘      要')]:
        ws.cell(h1,col).value = h
        ws.cell(h1,col).font  = Font(name='MS Gothic', size=9)
        ws.cell(h1,col).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(h1,col).border = thin_border()
    ws.row_dimensions[h1].height = 16
    r += 1
    # ヘッダー2行目（A-B結合で「種別」）
    h2 = r
    ws.merge_cells(f'A{h2}:B{h2}')
    ws.cell(h2,1).value = '種    別'
    ws.cell(h2,1).font  = Font(name='MS Gothic', size=9)
    ws.cell(h2,1).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(h2,1).border = thin_border()
    ws.merge_cells(f'C{h1}:C{h2}')
    ws.merge_cells(f'D{h1}:D{h2}')
    ws.merge_cells(f'E{h1}:E{h2}')
    ws.cell(h2,3).border = thin_border()
    ws.cell(h2,4).border = thin_border()
    ws.cell(h2,5).border = thin_border()
    ws.row_dimensions[h2].height = 16
    r += 1
    # 空行
    ws.merge_cells(f'B{r}:E{r}')
    for c in range(1, 6):
        ws.cell(r,c).border = thin_border()
    ws.row_dimensions[r].height = 25
    r += 1
    return r

def build_category_sheet(ws, rows, subtotal_map):
    ws.title = '工事別数量書'
    ws.column_dimensions['A'].width = 6.6
    ws.column_dimensions['B'].width = 40.5
    ws.column_dimensions['C'].width = 24.5
    ws.column_dimensions['D'].width = 22.6
    ws.column_dimensions['E'].width = 30.8

    data = [row for row in rows
            if len(row) >= 2 and not all(v == '' for v in row) and not is_page_num(row[0])]
    total_pages = (len(data) + CAT_ROWS_PER_PAGE - 1) // CAT_ROWS_PER_PAGE

    # ── ヘッダー（1ページ目） ──
    # タイトル（中央寄せ＋外枠罫線）
    r = 1
    ws.merge_cells(f'A{r}:E{r}')
    ws.cell(r,1).value = '工 種 別 数 量 書'
    ws.cell(r,1).font  = Font(name='MS Gothic', size=10, bold=True)
    ws.cell(r,1).alignment = Alignment(horizontal='center', vertical='center')
    for c in range(1, 6):
        ws.cell(r,c).border = thin_border()
    ws.row_dimensions[r].height = 20
    r += 1

    # ヘッダー1行目（A-B結合で「工種」、A-B間の縦線なし）
    h1 = r
    ws.merge_cells(f'A{h1}:B{h1}')
    ws.cell(h1,1).value = '工    種'
    ws.cell(h1,1).font  = Font(name='MS Gothic', size=9)
    ws.cell(h1,1).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(h1,1).border = thin_border()
    for col, h in [(3,'内  容 (数量)'),(4,'金　額 （円）'),(5,' 摘      要')]:
        ws.cell(h1,col).value = h
        ws.cell(h1,col).font  = Font(name='MS Gothic', size=9)
        ws.cell(h1,col).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(h1,col).border = thin_border()
    ws.row_dimensions[h1].height = 16
    r += 1

    # ヘッダー2行目（A-B結合で「種別」）
    h2 = r
    ws.merge_cells(f'A{h2}:B{h2}')
    ws.cell(h2,1).value = '種    別'
    ws.cell(h2,1).font  = Font(name='MS Gothic', size=9)
    ws.cell(h2,1).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(h2,1).border = thin_border()
    ws.merge_cells(f'C{h1}:C{h2}')
    ws.merge_cells(f'D{h1}:D{h2}')
    ws.merge_cells(f'E{h1}:E{h2}')
    ws.cell(h2,3).border = thin_border()
    ws.cell(h2,4).border = thin_border()
    ws.cell(h2,5).border = thin_border()
    ws.row_dimensions[h2].height = 16
    r += 1

    # ── データ行（ページ区切りあり：PDF忠実再現） ──
    current_big  = None
    section_rows = []   # 「計」(直接工事費小計)用：1-1〜1-n の金額行
    grand_rows   = []   # 「合 計」用：計＋共通仮設費＋現場管理費＋一般管理費等 の行
    page_num     = 1
    data_count   = 0

    for row in data:
        vals = (row + [''] * 5)[:5]
        num, name, qty, amt, note = vals
        bold        = bool(num)
        is_subtotal = name in SUBTOTAL_NAMES
        is_total    = name in {'合 計', '合　計', '合　　計'}

        # 工事全体名行（番号なし・名称あり・数量なし）＝見出し行：B~Eを結合
        if name and not num and not qty and not is_subtotal:
            apply_cell(ws.cell(r,1), '', center=True)
            apply_cell(ws.cell(r,2), name, bold=True)
            for c in range(3, 6):
                ws.cell(r,c).border = thin_border()
            ws.merge_cells(f'B{r}:E{r}')
            ws.row_dimensions[r].height = 25
            r += 1
            data_count += 1
            if data_count >= CAT_ROWS_PER_PAGE:
                r = write_cat_header_block(ws, r, page_num, total_pages)
                page_num  += 1
                data_count = 0
            continue

        if num.startswith('【'):
            current_big = num; section_rows = []; grand_rows = []

        apply_cell(ws.cell(r,1), num,  bold=bold, center=True)
        apply_cell(ws.cell(r,2), name, bold=bold or is_subtotal, center=is_subtotal)
        apply_cell(ws.cell(r,3), qty,  center=True)

        if qty == '一式' and num and not num.startswith('【') and not is_subtotal:
            # 直接工事費の細目（1-1等）→数量書の小計を参照
            ref = subtotal_map.get((current_big, num))
            if ref:
                set_formula(ws.cell(r,4), f'=数量書!G{ref}', right=True)
                section_rows.append(r)
            else:
                apply_cell(ws.cell(r,4), '', right=True)
            apply_cell(ws.cell(r,5), note)
            # 共通仮設費(2)/現場管理費(3)/一般管理費等(4)/有価発生材(5)は合計対象
            if '-' not in num:
                grand_rows.append(r)
        elif is_total:
            # 「合 計」＝ 計＋共通仮設費＋現場管理費＋…（改ページ行を挟むため加算式で連結）
            if grand_rows:
                set_formula(ws.cell(r,4),
                            '=' + '+'.join(f'D{x}' for x in grand_rows), right=True)
            else:
                apply_cell(ws.cell(r,4), '', right=True)
            apply_cell(ws.cell(r,5), '')
            grand_rows = []
        elif is_subtotal and section_rows:
            # 「計」＝ 直接工事費（1-1〜1-n）の合算（改ページ行を挟むため加算式で連結）
            formula = '=' + '+'.join(f'D{x}' for x in section_rows)
            set_formula(ws.cell(r,4), formula, right=True)
            apply_cell(ws.cell(r,5), '')
            grand_rows.append(r)  # 「計」も合計対象に含める
            section_rows = []
        else:
            apply_cell(ws.cell(r,4), '', right=True)
            apply_cell(ws.cell(r,5), note)

        ws.row_dimensions[r].height = 25  # ★ 全データ行25pt均一
        r += 1
        data_count += 1
        if data_count >= CAT_ROWS_PER_PAGE:
            r = write_cat_header_block(ws, r, page_num, total_pages)
            page_num  += 1
            data_count = 0

    # 罫線を一括補完（結合セル内側含む）。工事別のタイトルは枠内なので罫線あり
    finalize_borders(ws, 5)
    return ws

# ---------- 総括数量書 ----------

def build_summary_sheet(ws, rows, ws_cat):
    ws.title = '総括数量書'
    ws.column_dimensions['A'].width = 6.6
    ws.column_dimensions['B'].width = 25.7
    ws.column_dimensions['C'].width = 16.7
    ws.column_dimensions['D'].width = 25.7
    ws.column_dimensions['E'].width = 20.7
    ws.column_dimensions['F'].width = 20.7
    ws.column_dimensions['G'].width = 4.2

    # ── ヘッダー ──
    ws.merge_cells('A1:G1')
    ws['A1'].value = '総　括　数　量　書'
    ws['A1'].font  = Font(name='MS Gothic', size=11)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    # タイトル行の外枠罫線
    for c in range(1, 8):
        ws.cell(1,c).border = thin_border()
    ws.row_dimensions[1].height = 32

    ws.cell(2,1).border = thin_border()
    ws.merge_cells('B2:C2')
    ws.cell(2,2).value = '種     別'
    ws.cell(2,2).font  = Font(name='MS Gothic', size=9)
    ws.cell(2,2).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(2,2).border = thin_border()
    ws.cell(2,3).border = thin_border()
    ws.cell(2,4).value = '  内  容 (数量)'
    ws.cell(2,4).font  = Font(name='MS Gothic', size=9)
    ws.cell(2,4).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(2,4).border = thin_border()
    ws.cell(2,5).value = '金　額 （円）'
    ws.cell(2,5).font  = Font(name='MS Gothic', size=9)
    ws.cell(2,5).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(2,5).border = thin_border()
    ws.cell(2,7).border = thin_border()
    ws.row_dimensions[2].height = 16

    ws.cell(3,1).border = thin_border()
    ws.merge_cells('B3:C3')
    ws.cell(3,2).value = '内     訳'
    ws.cell(3,2).font  = Font(name='MS Gothic', size=9)
    ws.cell(3,2).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(3,2).border = thin_border()
    ws.cell(3,3).border = thin_border()
    ws.cell(3,4).border = thin_border()
    ws.cell(3,5).border = thin_border()
    ws.cell(3,6).border = thin_border()
    ws.cell(3,7).border = thin_border()
    ws.row_dimensions[3].height = 16

    ws.merge_cells('D2:D3')
    ws.merge_cells('E2:E3')
    ws.merge_cells('F2:G3')
    ws.cell(2,6).value = '摘　　要'
    ws.cell(2,6).font  = Font(name='MS Gothic', size=9)
    ws.cell(2,6).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(2,6).border = thin_border()

    # 工事別数量書の「合 計」行番号をマップ
    cat_total_map = {}
    big_code = None
    for i, row in enumerate(ws_cat.iter_rows(min_row=4, values_only=True), 4):
        if not row: continue
        code = str(row[0]).strip() if row[0] else ''
        name = str(row[1]).strip() if row[1] else ''
        if code.startswith('【'): big_code = code
        if name in {'合 計', '合　計', '合　　計'} and big_code:
            cat_total_map[big_code] = i

    def blank_row(r):
        """空行（罫線枠付き）を出力"""
        apply_cell(ws.cell(r,1), '', center=True)
        ws.merge_cells(f'B{r}:C{r}')
        ws.cell(r,2).border = thin_border()
        ws.cell(r,3).border = thin_border()
        ws.cell(r,4).border = thin_border()
        ws.cell(r,5).border = thin_border()
        ws.merge_cells(f'F{r}:G{r}')
        ws.cell(r,6).border = thin_border()
        ws.cell(r,7).border = thin_border()
        ws.row_dimensions[r].height = 25

    r = 4
    item_rows = []
    price_total_row = None  # 工事価格合計の行
    tax_row = None          # 消費税額の行
    data = [row for row in rows
            if len(row) >= 2 and not all(v == '' for v in row) and not is_page_num(row[0])]

    for vals in data:
        vals = (vals + [''] * 5)[:5]
        num, name, qty, amt, note = vals
        is_total_label = name in {'工事価格 合計', '工事価格　合計', '消費税額',
                                  '総 合 計', '総　合　計'}
        # 工事全体名行（番号なし・名称あり・数量なし、かつ合計ラベルでない）＝見出し：B~G結合
        is_proj_title = bool(name) and not num and not qty and not is_total_label

        if is_proj_title:
            apply_cell(ws.cell(r,1), '', center=True)
            apply_cell(ws.cell(r,2), name, bold=True)
            for c in range(3, 8):
                ws.cell(r,c).border = thin_border()
            ws.merge_cells(f'B{r}:G{r}')
            ws.row_dimensions[r].height = 25
            r += 1
            continue

        # 合計ラベルの前に空行を1行入れる（PDF様式再現）
        if name in {'工事価格 合計', '工事価格　合計', '消費税額', '総 合 計', '総　合　計'}:
            blank_row(r)
            r += 1

        bold = bool(num)
        is_center_label = name in {'総 合 計', '総　合　計'}
        apply_cell(ws.cell(r,1), num, bold=bold, center=True)
        ws.merge_cells(f'B{r}:C{r}')
        apply_cell(ws.cell(r,2), name, bold=bold, center=is_center_label)
        ws.cell(r,3).border = thin_border()
        apply_cell(ws.cell(r,4), qty, center=True)

        # 工事番号は【1】形式。そのままキーとして工事別の合計行を参照
        if num.startswith('【') and qty == '一式':
            ref = cat_total_map.get(num)
            if ref: set_formula(ws.cell(r,5), f'=工事別数量書!D{ref}', right=True)
            else:   apply_cell(ws.cell(r,5), '', right=True)
            item_rows.append(r)
        elif name in {'工事価格 合計', '工事価格　合計'} and item_rows:
            set_formula(ws.cell(r,5), f'=SUM(E{item_rows[0]}:E{item_rows[-1]})', right=True)
            price_total_row = r
        elif name == '消費税額' and price_total_row:
            set_formula(ws.cell(r,5), f'=INT(E{price_total_row}*0.1)', right=True)
            tax_row = r
        elif name in {'総 合 計', '総　合　計'} and price_total_row and tax_row:
            set_formula(ws.cell(r,5), f'=E{price_total_row}+E{tax_row}', right=True)
        else:
            apply_cell(ws.cell(r,5), '', right=True)

        ws.merge_cells(f'F{r}:G{r}')
        apply_cell(ws.cell(r,6), note)
        ws.cell(r,7).border = thin_border()
        # 結合セル内側にも罫線を補完
        for c in range(1, 8):
            if ws.cell(r,c).border.top is None or ws.cell(r,c).border.top.style is None:
                ws.cell(r,c).border = thin_border()
        ws.row_dimensions[r].height = 25
        r += 1

    # PDF様式に合わせ末尾に空行枠を追加（総合計の下に6行）
    for _ in range(6):
        blank_row(r)
        r += 1

    # 罫線を一括補完（結合セル内側含む）。総括のタイトルは枠内なので罫線あり
    finalize_borders(ws, 7)

# ---------- Excel生成 ----------

def build_excel(summary_rows, category_rows, quantity_rows):
    wb = openpyxl.Workbook()
    ws_quantity  = wb.active
    subtotal_map = build_quantity_sheet(ws_quantity, quantity_rows)
    ws_category  = wb.create_sheet()
    build_category_sheet(ws_category, category_rows, subtotal_map)
    ws_summary   = wb.create_sheet()
    build_summary_sheet(ws_summary, summary_rows, ws_category)
    wb.move_sheet(ws_summary, offset=-wb.index(ws_summary))
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# ============================================================
# Streamlit UI
# ============================================================

st.set_page_config(page_title='工事数量書 変換ツール', page_icon='📄', layout='centered')

st.title('📄 工事数量書 PDF → Excel 変換')
st.caption('JKK工事数量書のPDFをアップロードするとExcelに変換します')
st.divider()

uploaded = st.file_uploader('PDFファイルを選択', type='pdf', label_visibility='collapsed')

if uploaded:
    st.info(f'ファイル: **{uploaded.name}**　({uploaded.size / 1024:.0f} KB)')

    if st.button('変換開始', type='primary', use_container_width=True):
        progress = st.progress(0, text='準備中…')
        status   = st.empty()

        try:
            def cb(pct, msg):
                progress.progress(pct, text=msg)

            summary_rows, category_rows, quantity_rows = extract_all_pages(uploaded, cb)

            progress.progress(0.9, text='Excel生成中…')
            excel_buf = build_excel(summary_rows, category_rows, quantity_rows)

            progress.progress(1.0, text='完了！')
            status.success(
                f'変換完了 ✅　'
                f'総括: {len(summary_rows)}行 ／ '
                f'工事別: {len(category_rows)}行 ／ '
                f'数量: {len(quantity_rows)}行'
            )

            out_name = uploaded.name.replace('.pdf', '.xlsx').replace('.PDF', '.xlsx')
            st.download_button(
                label='📥 Excelをダウンロード',
                data=excel_buf,
                file_name=out_name,
                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                use_container_width=True,
                type='primary',
            )

        except Exception as e:
            progress.empty()
            st.error(f'エラーが発生しました: {e}')

st.divider()
st.caption('対応フォーマット: 総括数量書 / 工事別数量書 / 数量書 の3シート構成PDF')
